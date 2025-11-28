import openpyxl
from datetime import datetime
from utility.config_reader import get_config
import logging

# Logger for this file
logger = logging.getLogger(__name__)

class ExcelUtil:
    # Read tester name from config.ini
    tester = get_config("Excel", "tester")

    def __init__(self, excel_path,sheet):
        """
            Constructor to:
            Load Excel workbook into memory
            Select the correct sheet
            Store column mappings
        """
        logger.info(f"Initializing ExcelUtil with file: {excel_path}, sheet: {sheet}")

        # Store Excel file path
        self.excel_path = excel_path

        # Open workbook (load Excel file into memory)
        self.workbook = openpyxl.load_workbook(self.excel_path)

        # gets sheet that has data
        self.sheet= self.workbook[sheet]

        # Column mapping based on Excel structure
        self.COL_USERNAME = 7
        self.COL_PASSWORD = 8
        self.COL_EXPECTED = 9
        self.COL_TESTER = 3
        self.COL_DATE = 4
        self.COL_TIME = 5
        self.COL_RESULT = 10
        self.COL_ACTUAL = 11


     # gets data for Pytest parametrize
    def get_param_data(self):
         # Reads all rows from Excel (skipping header row) & Returns list of tuples: [(row,username, password), ...]
        logger.info("Reading parameterized test data from Excel")

        data_list = []
         # Start from row 2 (skip header row)
        for row in range(2, self.sheet.max_row + 1):  # Start from row 2, skip headers
            username = self.sheet.cell(row=row, column=self.COL_USERNAME).value or "" # Column Username
            password = self.sheet.cell(row=row, column=self.COL_PASSWORD).value or "" # Column Password
            error_message=self.sheet.cell(row=row,column=self.COL_EXPECTED).value or ""

            data_list.append((row, username, password,error_message)) # Append tuple (row_no, username, password,error_message)

        # Close workbook after reading
        self.workbook.close()
        logger.info("Workbook closed after reading")
        return data_list

    # Write test result
    def write_test_result(self,row, result,actual_output,tester):
        """
        Writes test execution results back to Excel:
        ✔ Tester
        ✔ Date
        ✔ Time
        ✔ Test Result (Pass / Fail)
        ✔ Actual Output/Error Message
        """

        logger.info(f"Writing test result for row {row}")

        now = datetime.now()
        self.sheet.cell(row=row, column=self.COL_TESTER).value = tester                   # Tester
        self.sheet.cell(row=row, column=self.COL_DATE).value = now.strftime("%Y-%m-%d")   # Date
        self.sheet.cell(row=row, column=self.COL_TIME).value = now.strftime("%H:%M:%S")   # Time
        self.sheet.cell(row=row, column=self.COL_RESULT).value = result                   # Result
        self.sheet.cell(row=row, column=self.COL_ACTUAL).value = actual_output            # actual output


        # Save changes back to Excel
        self.workbook.save(self.excel_path)
        logger.info("Excel file saved successfully after writing results")

        # Close workbook after writing
        self.workbook.close()
        logger.info("Workbook closed after writing")


    # Return only one specific row for a test
    def get_row(self, row_number):
        """
        Returns ONLY one specific row of data from Excel.
        Format → [(row_no, username, password)]
        """
        logger.info(f"Fetching specific row: {row_number}")
        row = row_number
        username = self.sheet.cell(row=row, column=self.COL_USERNAME).value
        password = self.sheet.cell(row=row, column=self.COL_PASSWORD).value
        return [(row, username, password)]
